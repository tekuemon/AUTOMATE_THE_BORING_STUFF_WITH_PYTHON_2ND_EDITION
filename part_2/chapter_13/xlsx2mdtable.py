# ライブラリのインストール
# pip install openpyxl

# ライブラリのインポート
import openpyxl, os

# xlsx...etcを取得
wb = openpyxl.load_workbook('世界各国基礎情報_正確性保証なし.xlsx')
sheet = wb['Sheet1']  # Sheet1 を取得
max_row = sheet.max_row  # 最大行
max_column = sheet.max_column  # 最大列

# 書き込み対象Markdown open
md = open('世界各国基礎情報_正確性保証なし.md', 'w')

# セルの値をMarkdownに形式を変えて転記
for i in range(2, max_row+1):
    # 見出しを国名とする
    md.write('# ' + str(sheet.cell(row=i, column=1).value + '\n'))
    # テーブルとして転記
    for j in range(2, max_column+1):
        # テーブルの頭作成
        if(j==2):
            md.write('|  |  |' + '\n')
            md.write('| --- | --- |' + '\n')
        # 各情報転記
        md.write('| ' + str(sheet.cell(row=1, column=j).value).replace('\xb2', '') +
                 ' | ' + str(sheet.cell(row=i, column=j).value).replace('\xb2', '') +
                 ' |' + '\n')
        # 各国区切り
        if(j==max_column):
            md.write('\n')

# 書き込み対象Markdown close
md.close()
